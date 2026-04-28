#!/usr/bin/env python3
"""
Excel Device Feature Parser - Command-line tool to parse and query device feature data.

This tool extracts hierarchical feature/sub-feature information from Excel workbooks
and allows searching for devices based on their supported features.
"""

import argparse
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional


class FeatureParser:
    """Parser for extracting and querying device feature data from Excel files."""
    
    def __init__(self, excel_path: str, sheet_name: str = "HPX Feature Table"):
        """
        Initialize the parser with an Excel file.
        
        Args:
            excel_path: Path to the Excel file
            sheet_name: Name of the sheet to parse (default: "HPX Feature Table")
        """
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.features = {}  # {main_feature: [sub_features]}
        self.feature_columns = {}  # {feature_name: column_index}
        self.devices = []  # List of device names
        self.data_start_row = 4  # Data starts at row 4 (after 3 header rows)
        self.feature_start_col = 10  # Features start at column 10
        
        self._load_workbook()
        self._parse_structure()
    
    def _load_workbook(self):
        """Load the Excel workbook and validate the sheet exists."""
        if not self.excel_path.exists():
            raise FileNotFoundError(
                f"Excel file not found: {self.excel_path}\n"
                f"Please provide a valid path to the Excel file."
            )
        
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")
        
        if self.sheet_name not in self.workbook.sheetnames:
            available = ", ".join(self.workbook.sheetnames)
            raise ValueError(
                f"Sheet '{self.sheet_name}' not found.\n"
                f"Available sheets: {available}"
            )
        
        self.worksheet = self.workbook[self.sheet_name]
    
    def _parse_structure(self):
        """Parse the Excel structure to extract features, sub-features, and devices."""
        ws = self.worksheet
        
        # Parse features from row 1 and row 2
        current_main_feature = None
        
        for col_idx in range(self.feature_start_col, ws.max_column + 1):
            # Get main feature from row 1
            main_feature_cell = ws.cell(1, col_idx).value
            if main_feature_cell and str(main_feature_cell).strip() and \
               str(main_feature_cell).strip() != "Main Feature":
                current_main_feature = str(main_feature_cell).strip()
                if current_main_feature not in self.features:
                    self.features[current_main_feature] = []
            
            # Get sub-feature from row 2
            sub_feature_cell = ws.cell(2, col_idx).value
            # Check if sub-feature exists and is not None, not "Sub-Feature", and not empty
            has_sub_feature = (sub_feature_cell is not None and 
                             str(sub_feature_cell).strip() and 
                             str(sub_feature_cell).strip() != "Sub-Feature")
            
            if has_sub_feature:
                sub_feature = str(sub_feature_cell).strip()
                
                # If no main feature yet, this IS a main feature (no sub-features)
                if current_main_feature is None:
                    # Check if row 1 has a value
                    row1_val = ws.cell(1, col_idx).value
                    if row1_val and str(row1_val).strip() != "Main Feature":
                        current_main_feature = str(row1_val).strip()
                        self.features[current_main_feature] = []
                    else:
                        # This column itself is a standalone feature
                        current_main_feature = sub_feature
                        self.features[current_main_feature] = []
                        self.feature_columns[current_main_feature] = col_idx
                        continue
                
                # Add sub-feature to current main feature
                if current_main_feature:
                    if sub_feature not in self.features[current_main_feature]:
                        self.features[current_main_feature].append(sub_feature)
                    self.feature_columns[sub_feature] = col_idx
            else:
                # No sub-feature (row 2 is None or empty)
                # The main feature itself is a standalone column
                if current_main_feature and current_main_feature not in self.feature_columns:
                    self.feature_columns[current_main_feature] = col_idx
        
        # Parse device names from column 4 (Platform column)
        for row_idx in range(self.data_start_row, ws.max_row + 1):
            device_cell = ws.cell(row_idx, 4).value
            if device_cell and str(device_cell).strip():
                self.devices.append(str(device_cell).strip())
    
    def _is_feature_supported(self, row_idx: int, col_idx: int) -> bool:
        """
        Check if a feature is supported for a device at given row and column.
        
        Args:
            row_idx: Row index in the worksheet
            col_idx: Column index in the worksheet
            
        Returns:
            True if feature is supported, False otherwise
        """
        cell_value = self.worksheet.cell(row_idx, col_idx).value
        
        if cell_value is None or cell_value == "":
            return False
        
        # Convert to string and check
        val_str = str(cell_value).strip().upper()
        
        # Check for various unsupported indicators
        unsupported_values = ["N/A", "NA", "N.A.", ""]
        if val_str in unsupported_values:
            return False
        
        # Any non-empty value that's not explicitly unsupported is considered supported
        # Common values: Y, X, ✓, V, 1, etc.
        return True
    
    def list_all_features(self) -> str:
        """
        Generate a formatted string listing all features and sub-features.
        
        Returns:
            Formatted string with hierarchical feature list
        """
        output = []
        output.append(f"\nParsing '{self.sheet_name}' sheet from {self.excel_path.name}...\n")
        
        total_main = len(self.features)
        total_sub = sum(len(subs) for subs in self.features.values())
        
        output.append(f"Found {total_main} main features with {total_sub} sub-features:\n")
        
        for main_feature, sub_features in self.features.items():
            output.append(f"\n{main_feature}")
            if sub_features:
                for idx, sub_feature in enumerate(sub_features):
                    prefix = "└─" if idx == len(sub_features) - 1 else "├─"
                    output.append(f"  {prefix} {sub_feature}")
            else:
                output.append("  (No sub-features)")
        
        return "\n".join(output)
    
    def search_by_feature(self, feature_query: str) -> str:
        """
        Search for devices that support a specific feature.
        
        Args:
            feature_query: Feature name to search for (partial match supported)
            
        Returns:
            Formatted string with search results
        """
        feature_query_lower = feature_query.lower()
        
        # Find matching features
        matching_features = []
        for main_feature, sub_features in self.features.items():
            if feature_query_lower in main_feature.lower():
                matching_features.append((main_feature, main_feature, True))  # (display_name, search_name, is_main)
            
            for sub_feature in sub_features:
                if feature_query_lower in sub_feature.lower():
                    matching_features.append((sub_feature, sub_feature, False))
        
        if not matching_features:
            return (f"\nNo features matching '{feature_query}' found.\n"
                   f"Use --list-features to see all available features.")
        
        # If multiple matches, show all
        if len(matching_features) > 1:
            output = [f"\nMultiple features matching '{feature_query}':\n"]
            for display_name, _, is_main in matching_features:
                output.append(f"  - {display_name}")
            output.append("\nPlease be more specific or use one of these exact names.")
            return "\n".join(output)
        
        # Single match - find supporting devices
        display_name, search_name, is_main = matching_features[0]
        
        # Find the feature category
        feature_category = None
        for main_feat, sub_feats in self.features.items():
            if search_name == main_feat or search_name in sub_feats:
                feature_category = main_feat
                break
        
        # Get the column index for this feature
        if search_name not in self.feature_columns:
            return f"\nFeature '{search_name}' found but column index not available."
        
        col_idx = self.feature_columns[search_name]
        
        # Find supporting devices
        supporting_devices = []
        not_supporting_devices = []
        
        for row_idx in range(self.data_start_row, self.worksheet.max_row + 1):
            device_name = self.worksheet.cell(row_idx, 4).value
            if device_name and str(device_name).strip():
                device_name = str(device_name).strip()
                if self._is_feature_supported(row_idx, col_idx):
                    supporting_devices.append(device_name)
                else:
                    not_supporting_devices.append(device_name)
        
        # Format output
        output = []
        output.append(f"\nSearching for: '{display_name}'")
        if feature_category:
            output.append(f"Feature Category: {feature_category}")
        output.append("")
        
        if supporting_devices:
            output.append(f"Devices supporting '{display_name}':")
            for idx, device in enumerate(supporting_devices, 1):
                output.append(f"  {idx}. {device}")
            output.append(f"\nTotal: {len(supporting_devices)} device(s) found")
        else:
            output.append(f"No devices support '{display_name}'")
        
        if not_supporting_devices:
            output.append(f"\nDevices NOT supporting '{display_name}': {len(not_supporting_devices)}")
        
        return "\n".join(output)
    
    def reverse_search(self, device_query: str) -> str:
        """
        Find all features supported by a specific device.
        
        Args:
            device_query: Device name to search for (partial match supported)
            
        Returns:
            Formatted string with device feature support
        """
        device_query_lower = device_query.lower()
        
        # Find matching devices
        matching_devices = []
        device_row_map = {}
        
        for row_idx in range(self.data_start_row, self.worksheet.max_row + 1):
            device_name = self.worksheet.cell(row_idx, 4).value
            if device_name and str(device_name).strip():
                device_name = str(device_name).strip()
                if device_query_lower in device_name.lower():
                    matching_devices.append(device_name)
                    device_row_map[device_name] = row_idx
        
        if not matching_devices:
            return (f"\nNo devices matching '{device_query}' found.\n"
                   f"Please check the device name and try again.")
        
        # If multiple matches, show all
        if len(matching_devices) > 1:
            output = [f"\nMultiple devices matching '{device_query}':\n"]
            for device in matching_devices:
                output.append(f"  - {device}")
            output.append("\nPlease be more specific or use one of these exact names.")
            return "\n".join(output)
        
        # Single match - show all features
        device_name = matching_devices[0]
        row_idx = device_row_map[device_name]
        
        output = []
        output.append(f"\nFeatures supported by '{device_name}':\n")
        
        total_supported = 0
        total_not_supported = 0
        
        for main_feature, sub_features in self.features.items():
            output.append(f"\n{main_feature}:")
            
            if sub_features:
                # Check each sub-feature
                for sub_feature in sub_features:
                    if sub_feature in self.feature_columns:
                        col_idx = self.feature_columns[sub_feature]
                        is_supported = self._is_feature_supported(row_idx, col_idx)
                        symbol = "✓" if is_supported else "✗"
                        output.append(f"  {symbol} {sub_feature}")
                        if is_supported:
                            total_supported += 1
                        else:
                            total_not_supported += 1
            else:
                # Main feature with no sub-features
                if main_feature in self.feature_columns:
                    col_idx = self.feature_columns[main_feature]
                    is_supported = self._is_feature_supported(row_idx, col_idx)
                    symbol = "✓" if is_supported else "✗"
                    output.append(f"  {symbol} (Main feature)")
                    if is_supported:
                        total_supported += 1
                    else:
                        total_not_supported += 1
        
        output.append(f"\n{'='*60}")
        output.append(f"Summary: {total_supported} features supported, {total_not_supported} not supported")
        
        return "\n".join(output)
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()


def main():
    """Main entry point for the command-line interface."""
    parser = argparse.ArgumentParser(
        description="Parse and query device feature data from Excel workbooks.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --list-features devices.xlsx
  %(prog)s --search "Audio Levels" devices.xlsx
  %(prog)s --reverse-search "Watson" devices.xlsx
  %(prog)s --search "HP Sure Click" --both  # Search both Consumer and CMIT files
  %(prog)s --search "Audio Levels" --consumer  # Search Consumer file only
  %(prog)s --search "HP Sure Click" --commercial  # Search Commercial file only
  %(prog)s --sheet "Custom Sheet" --list-features devices.xlsx
        """
    )
    
    parser.add_argument(
        "excel_file",
        nargs="?",
        help="Path to the Excel file containing device feature data (optional if --both, --consumer, or --commercial is used)"
    )
    
    parser.add_argument(
        "--both",
        action="store_true",
        help="Search both Consumer and CMIT files automatically"
    )
    
    parser.add_argument(
        "--consumer",
        action="store_true",
        help="Search the Consumer file only"
    )
    
    parser.add_argument(
        "--commercial",
        action="store_true",
        help="Search the Commercial (CMIT) file only"
    )
    
    parser.add_argument(
        "--list-features", "-lf",
        action="store_true",
        help="List all features and sub-features in hierarchical format"
    )
    
    parser.add_argument(
        "--search", "-s",
        metavar="FEATURE",
        help="Search for devices supporting a specific feature"
    )
    
    parser.add_argument(
        "--reverse-search", "-rs",
        metavar="DEVICE",
        help="Find all features supported by a specific device"
    )
    
    parser.add_argument(
        "--sheet",
        default="HPX Feature Table",
        help="Name of the sheet to parse (default: HPX Feature Table)"
    )
    
    args = parser.parse_args()
    
    # Validate that at least one action is specified
    if not any([args.list_features, args.search, args.reverse_search]):
        parser.error("Please specify an action: --list-features, --search, or --reverse-search")
    
    # Validate file selection flags
    file_flags = sum([args.both, args.consumer, args.commercial, bool(args.excel_file)])
    if file_flags == 0:
        parser.error("Please specify a file: excel_file, --both, --consumer, or --commercial")
    if file_flags > 1:
        parser.error("Please specify only one of: excel_file, --both, --consumer, or --commercial")
    
    # Determine which files to process
    files_to_process = []
    
    # Get the directory where this script is located (device-data folder)
    script_dir = Path(__file__).parent
    consumer_file = script_dir / "HPX Release Matrix for Consumer_20241125.xlsx"
    cmit_file = script_dir / "myHP_HPX Release Matrix for CMIT.xlsx"
    
    if args.both:
        # Use both default files
        if consumer_file.exists():
            files_to_process.append(("Consumer", str(consumer_file)))
        if cmit_file.exists():
            files_to_process.append(("Commercial", str(cmit_file)))
        
        if not files_to_process:
            print("\nError: Could not find Consumer or CMIT Excel files in device-data directory.", file=sys.stderr)
            sys.exit(1)
    elif args.consumer:
        if not consumer_file.exists():
            print(f"\nError: Consumer file not found: {consumer_file}", file=sys.stderr)
            sys.exit(1)
        files_to_process.append(("Consumer", str(consumer_file)))
    elif args.commercial:
        if not cmit_file.exists():
            print(f"\nError: Commercial file not found: {cmit_file}", file=sys.stderr)
            sys.exit(1)
        files_to_process.append(("Commercial", str(cmit_file)))
    else:
        files_to_process.append(("", args.excel_file))
    
    # Process each file
    for file_type, excel_file in files_to_process:
        if len(files_to_process) > 1 or file_type:
            print("\n" + "="*70)
            print(f"{file_type} Devices")
            print("="*70)
        
        # Initialize parser
        try:
            feature_parser = FeatureParser(excel_file, args.sheet)
        except (FileNotFoundError, ValueError) as e:
            print(f"\nError processing {excel_file}: {e}", file=sys.stderr)
            continue
        
        try:
            # Execute requested action
            if args.list_features:
                print(feature_parser.list_all_features())
            
            if args.search:
                print(feature_parser.search_by_feature(args.search))
            
            if args.reverse_search:
                print(feature_parser.reverse_search(args.reverse_search))
        
        finally:
            feature_parser.close()


if __name__ == "__main__":
    main()
